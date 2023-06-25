import requests
import json
import datetime as dt
import pandas as pd
import yfinance as yf
import numpy as np
import pandas_ta as ta
import pyupbit
import math
import ccxt
import openpyxl
import time



access = 
secret = 
faccess = 
fsecret = 
upbit = pyupbit.Upbit(access,secret)

ftx = ccxt.ftx(
{
    'apiKey' : faccess,
    'secret' : fsecret
})

utest = ccxt.upbit(
{
    'apiKey' : access,
    'secret' : secret
})

############################################################################################## FUNCTIONS START 

def truncate(number, digits) -> float:
    nbDecimals = len(str(number).split('.')[1]) 
    if nbDecimals <= digits:
        return number
    stepper = 10.0 ** digits
    return math.trunc(stepper * number) / stepper

def get_data_duration(symbol,dur):    
    bars = ftx.fetch_ohlcv(symbol,timeframe = '1d',limit = dur)
    df = pd.DataFrame(bars,columns = ['Date Time','Open','High','Low','Close','Volume'])
    df['Date Time'] = pd.to_datetime(df['Date Time'],unit= 'ms')+dt.timedelta(hours = 9)
    df.index = df['Date Time']
    df = df.iloc[:,1:-1]
    df.iat[-1,-1] = ftx.fetch_order_book(symbol)['bids'][0][0]
    return df
# Return True if the symbol is in position to enter
def get_usd_krw_enter(symbol,std):
    dur = 5
    df1 = pyupbit.get_ohlcv("KRW-"+symbol,count = dur)
    df1.iloc[4]["close"] = pyupbit.get_orderbook("KRW-"+symbol)['orderbook_units'][0]["ask_price"]
    df2 = get_data_duration(symbol+"-PERP",dur)
    df= df2["Close"].div(df1['close'])*1304
    mean = (df[0]+df[1]+df[2]+df[3]+df[4])/5
    std = np.std([df[0],df[1],df[2],df[3],df[4]]) *std
    if std<0.006:
        std = 0.006
    if std>0.01:
        std = 0.01
    std += 0.0015
    bh = mean+std
    print("Checking Enter Point For : | ",symbol," | Bollinger High: |",bh,"| Current Decimal: |",df[4],"|")
    return bool(bh<=df[4])
# Return True if the symbol is in position to close
def get_usd_krw_close(symbol,std):
    dur = 5 
    df1 = pyupbit.get_ohlcv("KRW-"+symbol,count = dur)
    df1.iloc[4]["close"] = pyupbit.get_orderbook("KRW-"+symbol)['orderbook_units'][0]["bid_price"]
    df2 = get_data_duration(symbol+"-PERP",dur)
    df= df2["Close"].div(df1['close'])*1304
    mean = (df[0]+df[1]+df[2]+df[3]+df[4])/5
    std = np.std([df[0],df[1],df[2],df[3],df[4]]) *std
    if std<0.006:
        std = 0.006
    if std>0.01:
        std = 0.01
    std += 0.0015
    bl = mean - std
    print("Checking Close Point For : | ",symbol," | Bollinger Low: |",bl,"| Current Decimal: |",df[4],"|")

    return bool(df[4]<=bl)
# Return positive for Up->FTX transfer, negative for the opposite
def ku_balance(krw,usd,forex):
    usd_k = usd/forex
    mean = (2*krw+2*usd_k)/3
    if krw>2*usd_k:
        if 2*usd_k<mean*0.9:
            send = krw - mean
            if send/2>usd_k:
                send = usd_k*2*0.98
            
            if send > 950000:
                send = 950000

            return send*0.98
            
        else: return 0
    else:
        if krw<mean*0.9:

            send = usd_k - mean/2
            if send > 950000:
                send = 950000
            return -send*forex*0.98
        else:
            return 0

## Convert USD to Coin Amount with precision
def usd_precision(symbol,usd):
    a = usd/ftx.fetch_order_book(symbol)['bids'][0][0]
    return ftx.amountToPrecision(symbol, a)

# Balance Out FTX and UPBIT
def balance_out():
    global ftx_cur_deposit
    global upbit_cur_deposit
    global cur_trans_coin
    global ftx_amount_save
    global recent_ftx_deposit
    global recent_upbit_deposit
    global cur_trans_amount
    df = yf.download(tickers = 'KRWUSD=X' ,period ='1d')
    forex = df.iloc[-1]["Close"]
    send = ku_balance(upbit.get_balance(ticker = "KRW"),ftx.fetch_balance()['USD']['free'],forex)
    if -4<send<5050:
        send = 0
    if send>0:
        ticker = upbit_deposit_key[upbit_cur_deposit]
        upbit_cur_deposit = (upbit_cur_deposit+1)%2
        address = ftx_deposit_address[ticker]
        if address != ftx.fetchDepositAddress(ticker)['address']:
            return -1
        cur_trans_coin = ticker
        krw = send
        usd = send*forex
        upbit.buy_market_order("KRW-"+ticker, krw)
        a = ftx.create_market_sell_order(ticker+"-PERP",usd_precision(ticker+"-PERP",usd))
        ftx_amount_save[ticker] = float(a['info']['size'])
        time.sleep(5)
        if ftx.fetchDeposits():
            recent_ftx_deposit = ftx.fetchDeposits()[-1]
        else:
            recent_ftx_deposit = None
        b = utest.withdraw(ticker, truncate(utest.fetch_balance()[ticker]['free'],6)-1, address)
        if 'size' in b.keys():
            cur_trans_amount = b['size']
        else:
            cur_trans_amount = b['amount']
        return 2
    elif send<0:
        ticker = ftx_deposit_key[ftx_cur_deposit]
        ftx_cur_deposit = (ftx_cur_deposit+1)%2
        address = upbit_deposit_address[ticker]
        if address != utest.fetchDepositAddress(ticker)['address']:
            return -1
        cur_trans_coin = ticker
        usd = -send
        ftx.create_market_buy_order(ticker+"/USD",usd_precision(ticker+"/USD",usd))
        a = ftx.create_market_sell_order(ticker+"-PERP",usd_precision(ticker+"-PERP",usd))
        ftx_amount_save[ticker] = float(a['info']['size'])
        time.sleep(5)
        if utest.fetchDeposits():
            recent_upbit_deposit = utest.fetchDeposits()[-1]
        else:
            recent_upbit_deposit = None
        b = ftx.withdraw(ticker, ftx.fetch_balance()[ticker]['free'], address)
        if 'size' in b.keys():
            cur_trans_amount = b['size']
        else:
            cur_trans_amount = b['amount']
        return 3
    
    else:
        return 1

## Return KRW,USD that is available for order
def order_amount_ku():
    df = yf.download(tickers = 'KRWUSD=X' ,period ='1d')
    forex = df.iloc[-1]["Close"]
    krw = upbit.get_balance(ticker = "KRW")
    usd_k = ftx.fetch_balance()['USD']['free']/forex*2
    krw = min(krw,usd_k) * 0.99
    usd = krw*forex
    return krw,usd
    
############################################################################################# FUNCTIONS END

############################################################################################# VARIABLES START
ftx_deposit_address = {"TRX" : "","ALGO" : ""}
upbit_deposit_address = {"TRX" : "","ALGO" : ""}
ftx_deposit_key = ["TRX","ALGO"]
upbit_deposit_key = ["TRX","ALGO"]

ftx_amount_save = {"TRX": 0,"ALGO":0,"NEAR":0,"ETC":0,"AVAX":0,"WAVES":0,"DOGE":0}

enter_check = ["NEAR","ETC","AVAX","DOGE"]
close_check = []

buyable_pos_num = 2

ftx_cur_deposit = 0
upbit_cur_deposit = 0

recent_ftx_deposit = None
recent_upbit_deposit = None

cur_trans_coin = None
cur_trans_amount = None

n = 0
std = 1.3
############################################################################################# VARIABLES END



############################################################################################# MAIN CODE START

while True:
    try:
        ## Check Close
        for i,tick in enumerate(close_check):
            if get_usd_krw_close(tick,std):
                buyable_pos_num+=1
                ftx.create_market_buy_order(tick+"-PERP",ftx_amount_save[tick])
                ftx_amount_save[tick] = 0
                upbit.sell_market_order("KRW-"+tick, upbit.get_balance(ticker = tick))
                enter_check.append(close_check.pop(i))
                workbook = openpyxl.load_workbook("history.xlsx")
                sheet = workbook['Sheet1']
                row = sheet.max_row+1
                sheet[f"A{row}"].value = tick
                sheet[f"B{row}"].value = str(dt.datetime.now())
                sheet[f"C{row}"].value = pyupbit.get_orderbook("KRW-"+tick)['orderbook_units'][0]["bid_price"]
                sheet[f"D{row}"].value = ftx.fetch_order_book(tick+"-PERP")['bids'][0][0]
                sheet[f"E{row}"].value = "--"
                sheet[f"F{row}"].value = "CLOSE"
                workbook.save("history.xlsx")
                workbook.close()

                if n == 1:
                    n = 0
                time.sleep(10)
                break

                
        if n == 0:
            print("Balancing two exchange markets")
            n = balance_out()
            if n == -1:
                break

        ## Check Enter
        elif n == 1 and buyable_pos_num>0:
            print("Checking Enter Point...")
            for i,ticker in enumerate(enter_check):
                if get_usd_krw_enter(ticker,std):
                    krw,usd = order_amount_ku()
                    krw = krw/buyable_pos_num
                    usd = usd/buyable_pos_num
                    upbit.buy_market_order("KRW-"+ticker, krw)
                    a = ftx.create_market_sell_order(ticker+"-PERP",usd_precision(ticker+"-PERP",usd))
                    ftx_amount_save[ticker] = float(a['info']['size'])
                    close_check.append(enter_check.pop(i))
                    buyable_pos_num -= 1
                    workbook = openpyxl.load_workbook("history.xlsx")
                    sheet = workbook['Sheet1']
                    row = sheet.max_row+1
                    sheet[f"A{row}"].value = ticker
                    sheet[f"B{row}"].value = str(dt.datetime.now())
                    sheet[f"C{row}"].value = pyupbit.get_orderbook("KRW-"+ticker)['orderbook_units'][0]["ask_price"]
                    sheet[f"D{row}"].value = ftx.fetch_order_book(ticker+"-PERP")['bids'][0][0]
                    sheet[f"E{row}"].value = krw
                    sheet[f"F{row}"].value = "ENTER"
                    workbook.save("history.xlsx")
                    workbook.close()
                    break
        ## Check for withdrawal from Upbit to FTX
        elif n == 2:
            print("Checking For withdrawal from Upbit to FTX")
            if ftx.fetchDeposits():
                if recent_ftx_deposit != ftx.fetchDeposits()[-1] and ftx.fetchDeposits()[-1]['info']['coin'] == cur_trans_coin and ftx.fetch_balance()[cur_trans_coin]['free']>=cur_trans_amount*0.95:
                    ftx.create_market_buy_order(cur_trans_coin+"-PERP",ftx_amount_save[cur_trans_coin])
                    n = 4

        ## Check for withdrawal from FTX to Upbit
        elif n == 3:
            print("Checking for withdrawal from FTX to Upbit")
            if utest.fetchDeposits():
                if recent_upbit_deposit != utest.fetchDeposits()[-1] and utest.fetchDeposits()[-1]['info']['currency'] == cur_trans_coin and utest.fetch_balance()[cur_trans_coin]['free']>=cur_trans_amount*0.95:
                    ftx.create_market_buy_order(cur_trans_coin+"-PERP",ftx_amount_save[cur_trans_coin])
                    n = 5
        ## Try to sell deposited coin in FTX
        elif n == 4:
            print("Selling transfer coin in FTX")
            n = 0
            try:
                ftx.create_market_sell_order(cur_trans_coin+"/USD",ftx.fetch_balance()[cur_trans_coin]['free'])
            except:
                n = 4

        ## Try to sell deposited coin in Upbit
        elif n == 5:
            print("Selling transfer coin in Upbit")
            n = 0
            try:
                upbit.sell_market_order("KRW-"+cur_trans_coin, upbit.get_balance(ticker = cur_trans_coin))
            except:
                n = 5

        time.sleep(1.5)

        print("Current Positions: ",close_check)
        print("-------------------------------------------------------------------------------------------------------------------------")
        print(" ")

    except:
        workbook = openpyxl.load_workbook("history.xlsx")
        sheet = workbook['Sheet1']
        sheet[f"H{sheet.max_row+1}"].value = n
        

############################################################################################# MAIN CODE END